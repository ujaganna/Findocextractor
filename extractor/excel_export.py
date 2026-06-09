import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SEVERITY_COLORS = {
    "CRITICAL": "C00000",
    "HIGH":     "E26B0A",
    "MEDIUM":   "C09000",
    "LOW":      "538135",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")


def _style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_excel_report(metrics: dict, anomalies: list[dict]) -> bytes:
    """Generate a formatted 3-sheet Excel report: Summary, Anomaly Flags, Raw Metrics."""
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 28

    ws["A1"] = "FinDoc Extractor — Filing Summary"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Company", metrics.get("company_name")),
        ("Filing Type", metrics.get("filing_type")),
        ("Fiscal Year", metrics.get("fiscal_year")),
        ("Revenue (Current, $M)", metrics.get("revenue_current")),
        ("Revenue (Prior, $M)", metrics.get("revenue_prior")),
        ("Net Income (Current, $M)", metrics.get("net_income_current")),
        ("Net Income (Prior, $M)", metrics.get("net_income_prior")),
        ("EPS Basic", metrics.get("eps_basic")),
        ("EPS Diluted", metrics.get("eps_diluted")),
        ("EBITDA ($M)", metrics.get("ebitda")),
        ("Total Debt ($M)", metrics.get("total_debt")),
        ("Total Equity ($M)", metrics.get("total_equity")),
        ("Operating Cash Flow ($M)", metrics.get("operating_cash_flow")),
        ("Audit Opinion", metrics.get("audit_opinion")),
        ("Auditor", metrics.get("auditor_name")),
        ("Anomaly Flags Found", len(anomalies)),
    ]

    for i, (label, value) in enumerate(summary_rows, start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True)
        if i % 2 == 0:
            label_cell.fill = ALT_ROW_FILL
            ws.cell(row=i, column=2).fill = ALT_ROW_FILL
        ws.cell(row=i, column=2, value=value if value is not None else "N/A")

    # ── Sheet 2: Anomaly Flags ────────────────────────────────────────────────
    wa = wb.create_sheet("Anomaly Flags")
    wa.column_dimensions["A"].width = 36
    wa.column_dimensions["B"].width = 12
    wa.column_dimensions["C"].width = 60
    wa.row_dimensions[1].height = 22

    for col, header in enumerate(["Flag", "Severity", "Detail"], start=1):
        wa.cell(row=1, column=col, value=header)
    _style_header_row(wa, 1, 3)

    for i, flag in enumerate(anomalies, start=2):
        hex_color = SEVERITY_COLORS.get(flag["severity"], "AAAAAA")
        sev_fill = PatternFill("solid", fgColor=hex_color)
        wa.cell(row=i, column=1, value=flag["flag"])
        sev_cell = wa.cell(row=i, column=2, value=flag["severity"])
        sev_cell.fill = sev_fill
        sev_cell.font = Font(bold=True, color="FFFFFF")
        sev_cell.alignment = Alignment(horizontal="center")
        wa.cell(row=i, column=3, value=flag["detail"])

    if not anomalies:
        wa.cell(row=2, column=1, value="No anomalies detected.")
        wa.cell(row=2, column=1).font = Font(italic=True, color="538135")

    # ── Sheet 3: Raw Metrics ──────────────────────────────────────────────────
    wr = wb.create_sheet("Raw Metrics")
    wr.column_dimensions["A"].width = 36
    wr.column_dimensions["B"].width = 28
    for col, header in enumerate(["Metric", "Value"], start=1):
        wr.cell(row=1, column=col, value=header)
    _style_header_row(wr, 1, 2)

    for i, (k, v) in enumerate(metrics.items(), start=2):
        wr.cell(row=i, column=1, value=k).font = Font(bold=True)
        wr.cell(row=i, column=2, value=str(v) if v is not None else "N/A")
        if i % 2 == 0:
            wr.cell(row=i, column=1).fill = ALT_ROW_FILL
            wr.cell(row=i, column=2).fill = ALT_ROW_FILL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
